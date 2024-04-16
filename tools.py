
import pandas as pd
from openpyxl import load_workbook, Workbook
from collections import Counter


class tools:

    @staticmethod
    def build_adj_mat(nodes, edges):
    
        adj_mat = pd.DataFrame(0, columns=nodes, index = nodes)
        for itr_edge in edges:

            adj_mat.loc[itr_edge[0], itr_edge[1]] = 1

        return adj_mat


    @staticmethod
    def build_actual_structure(latent_adj_mat, present_sensors):

        actual_structure = []
        for itr_sensor in present_sensors:

            sensor_queue = latent_adj_mat.columns[latent_adj_mat.loc[itr_sensor] == 1].to_list()
            while sensor_queue:

                next_sensor = sensor_queue.pop(0)
                if next_sensor == itr_sensor:

                    continue

                if next_sensor in present_sensors:

                    actual_structure.append((itr_sensor, next_sensor))

                else:
                    sensor_queue.extend(latent_adj_mat.columns[latent_adj_mat.loc[next_sensor] == 1].to_list())


        return actual_structure
    

    @staticmethod
    def compare_graphs(files, store_location):

        graph_store = {}
        for itr, itr_graph in enumerate(files):

            filepath = rf'{store_location}\{itr_graph}.xlsx'
            nodes, edges = tools.read_graph_excel(filepath)
            graph_store[itr] = {}
            graph_store[itr]['nodes'] = nodes 
            graph_store[itr]['edges'] = edges 

        comparison = {}
        for itr, itr_graph in graph_store.items():

            for itr_comp in range(itr+1, len(files)):
            
                comparison = {'equivalent': [],
                                'not in 2' : [],
                                'not in 1' : []}

                for itr_edge in graph_store[itr]['edges']:

                    if itr_edge in graph_store[itr_comp]['edges']:

                        comparison['equivalent'].append(itr_edge)

                    elif itr_edge not in graph_store[itr_comp]['edges']:

                        comparison['not in 2'].append(itr_edge)

                for itr_edge in graph_store[itr_comp]['edges']:

                    if itr_edge not in itr_graph['edges']:

                        comparison['not in 1'].append(itr_edge)

                print('-------------------------------------------------------------------------------------------------')
                print(f'Comparison - {itr} and {itr_comp}')
                print('-------------------------------------------------------------------------------------------------')
                print(f'Edges present in both: {comparison["equivalent"]}')
                print('*************************************************************************************************')
                print(f'Edges not in {itr}: {comparison["not in 1"]}')
                print('*************************************************************************************************')
                print(f'Edges not in {itr_comp}: {comparison["not in 2"]}')
                print('*************************************************************************************************')
            


    @staticmethod
    def read_graph_excel(filepath):

        wb = load_workbook(filepath)
        ws = wb.active

        nodes = []
        edges = []

        for row in ws.iter_rows(values_only=True):

            if row[0]:

                nodes.append(row[0]) 

            if row[1] and row[2]:
                
                edges.append((row[1], row[2]))  

        return nodes, edges


    def print_to_excel(graph, save_path):

        wb = Workbook()
        ws = wb.active

        for row_index, row_data in enumerate(graph['nodes'], start=1):

            ws.cell(row=row_index, column=1, value=row_data)

        for row_index, row_data in enumerate(graph['edges'], start=1):
            for col_index, value in enumerate(row_data, start=2):
                ws.cell(row=row_index, column=col_index, value=value)

        wb.save(save_path)


    def compare_parallel_graphs(graph_set1, graph_set2, set1_path, set2_path):

        for itr in range(len(graph_set1)):

            graph1_file = graph_set1[itr]
            graph1_path = f'{set1_path}\{graph1_file}.xlsx'
            graph2_file = graph_set2[itr]
            graph2_path = f'{set2_path}\{graph2_file}.xlsx'

            graph1 = {}
            graph1['nodes'], graph1['edges'] = tools.read_graph_excel(graph1_path)
            graph2 = {}
            graph2['nodes'], graph2['edges'] = tools.read_graph_excel(graph2_path)

            graph1['undirected'] = []
            graph1['directed'] = []
            for itr_edge in graph1['edges']:

                if (itr_edge[1], itr_edge[0]) in graph1['edges'] and (itr_edge[1], itr_edge[0]) not in graph1['undirected']:
                    
                    graph1['undirected'].append(itr_edge)

                elif (itr_edge[1], itr_edge[0]) not in graph1['edges']:

                    graph1['directed'].append(itr_edge)

            graph2['undirected'] = []
            graph2['directed'] = []
            for itr_edge in graph2['edges']:

                if (itr_edge[1], itr_edge[0]) in graph2['edges'] and (itr_edge[1], itr_edge[0]) not in graph2['undirected']:
                    
                    graph2['undirected'].append(itr_edge)

                elif (itr_edge[1], itr_edge[0]) not in graph2['edges']:

                    graph2['directed'].append(itr_edge)

            
            # Comparison
            similarities = {}
            differences = {}
            all_diedges = graph1['directed'] + graph2['directed']
            counts = Counter(all_diedges)
            similarities['directed'] = [edge for edge, count in counts.items() if count == 2]
            differences['directed'] = [edge for edge, count in counts.items() if count == 1]

            all_undiedges = graph1['undirected'] + graph2['undirected']
            counts = Counter(all_undiedges)
            similarities['undirected'] = [edge for edge, count in counts.items() if count == 2]
            differences['undirected'] = [edge for edge, count in counts.items() if count == 1]

            differences['di_graph1'] = []
            differences['di_graph2'] = []
            for itr_di_diff in differences['directed']:

                if itr_di_diff in graph1['directed']:

                    differences['di_graph1'].append(itr_di_diff)

                elif itr_di_diff in graph2['directed']:

                    differences['di_graph2'].append(itr_di_diff)

                else:

                    raise ValueError('Edge in neither graphs.')

            differences['undi_graph1'] = []
            differences['undi_graph2'] = []
            for itr_undi_diff in differences['undirected']:

                if itr_undi_diff in graph1['undirected']:

                    differences['undi_graph1'].append(itr_undi_diff)

                elif itr_undi_diff in graph2['undirected']:

                    differences['undi_graph2'].append(itr_undi_diff)

                else:

                    raise ValueError('Edge in neither graphs.')
                

            print('-------------------------------------------------------------------------------------------------')
            print(f'Comparison - {graph1_file} and {graph2_file}')
            print('-------------------------------------------------------------------------------------------------')
            print(f'Directed similarities: {similarities["directed"]}')
            print('*************************************************************************************************')
            print(f'Undirected similarities: {similarities["undirected"]}')
            print('*************************************************************************************************')
            print(f'Additional directed edges in graph1: {differences["di_graph1"]}')
            print('*************************************************************************************************')
            print(f'Additional undirected edges in graph1: {differences["undi_graph1"]}')
            print('*************************************************************************************************')
            print(f'Additional directed edges in graph2: {differences["di_graph2"]}')
            print('*************************************************************************************************')
            print(f'Additional undirected edges in graph2: {differences["undi_graph2"]}')
            print('*************************************************************************************************')