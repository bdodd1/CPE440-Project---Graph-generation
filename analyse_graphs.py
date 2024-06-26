import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from itertools import product
from handle_cycles import handle_cycles
from test_run_fges import run_fges
import numpy as np
import ast
from scipy import stats 


class analyse_graphs:


    @staticmethod
    def kn_score_vs_graph_score(data_store):

        kn_scores = []
        graph_scores = []
        for itr_comb in data_store:

            graph_scores.append(itr_comb['graph score'])
            kn_scores.append(itr_comb['knowledge score'])

        plt.figure()
        plt.scatter(kn_scores, graph_scores)
        plt.xlabel('knowledge scores')
        plt.ylabel('graph scores')
        plt.show()


    @staticmethod
    def amount_kn_vs_graph_score(data_store):

        kn_amount = []
        graph_scores = []
        for itr_comb in data_store:

            graph_scores.append(itr_comb['graph score'])
            kn_amount.append(itr_comb['knowledge count'])

        plt.figure()
        plt.scatter(kn_amount, graph_scores)
        plt.xlabel('amount of knowledge')
        plt.ylabel('graph scores')
        plt.show()


    @staticmethod
    def pc_kn_vs_graph_score(data_store):

        pc_kn = []
        graph_scores = []
        for itr_comb in data_store:

            graph_scores.append(itr_comb['graph score'])
            pc_kn.append(itr_comb['pc knowledge'])

        plt.figure()
        plt.scatter(pc_kn, graph_scores)
        plt.xlabel('pc of final edges knowledge')
        plt.ylabel('graph scores')
        plt.show()


    @staticmethod
    def pc_max_degree_vs_graph_score(data_store):

        pc_max_degree = []
        graph_scores = []
        for itr_comb in data_store:

            graph_scores.append(itr_comb['graph score'])
            pc_max_degree.append(itr_comb[r'% of max degree'])

        plt.figure()
        plt.scatter(pc_max_degree, graph_scores)
        plt.xlabel(r'% of max degree')
        plt.ylabel('graph scores')
        plt.show()



    @staticmethod
    def graph_score_for_mode(data_store, test):

        test_modes_list = []
        for itr_comb in data_store:

            test_modes_list.append(ast.literal_eval(itr_comb['combination'])[test])

        test_modes = list(set(test_modes_list))
        mode_data_store = []
        for itr_mode in test_modes:

            mode_data_store.append([])

        for itr_comb in data_store:

            mode = ast.literal_eval(itr_comb['combination'])[test]
            mode_data_store[mode].append(itr_comb['graph score'])

        plt.figure()
        for itr_mode, itr_mode_data in enumerate(mode_data_store):

            plt.scatter([itr_mode]*len(itr_mode_data), itr_mode_data)

        plt.xlabel('mode')
        plt.ylabel('graph scores')
        plt.show()


    @staticmethod
    def avg_graph_score_vs_mode_bar(data_store, test):

        test_modes_list = []
        for itr_comb in data_store:

            test_modes_list.append(ast.literal_eval(itr_comb['combination'])[test])

        test_modes = list(set(test_modes_list))
        mode_data_store = []
        for itr_mode in test_modes:

            mode_data_store.append([])

        for itr_comb in data_store:

            mode = ast.literal_eval(itr_comb['combination'])[test]
            mode_data_store[mode].append(itr_comb['graph score'])

        data = []
        plt.figure()
        for itr_mode, itr_mode_data in enumerate(mode_data_store):

            data.append(sum(itr_mode_data)/len(itr_mode_data))
            plt.bar(itr_mode, sum(itr_mode_data)/len(itr_mode_data))

        if all(x<0 for x in data):
            plt.ylim(1.1*min(data), 0.9*max(data))
        elif any(x<0 for x in data) and any(x>0 for x in data):
            plt.ylim(1.1*min(data), 1.1*max(data))
        elif all(x>0 for x in data):
            plt.ylim(0.9*min(data), 1.1*max(data))
        plt.xlabel('mode')
        plt.ylabel('avg graph score')
        plt.show()

        print(100*(max(data)-min(data))/(0.5*max(data)+min(data)))
    

    @staticmethod
    def find_pc_unique_graphs(data_store):

        scores = []
        for itr_comb in data_store:

            scores.append(itr_comb['graph score'])

        return 100*len(set(scores)) / len(scores)


    @staticmethod
    def find_top_three_graphs(data_store):

        scores = []
        for itr_comb in data_store:

            scores.append(itr_comb['graph score'])

        unique_scores = list(set(scores))
        if len(unique_scores) < 3:
            top_score_count = len(unique_scores)
        else:
            top_score_count = 3
        top_three_scores = sorted(unique_scores, reverse=True)[:top_score_count]

        top_three_store = []
        for _ in range(len(top_three_scores)):
            top_three_store.append([])


        for itr_comb, itr_score in enumerate(scores):

            if itr_score == top_three_scores[0]:

                top_three_store[0].append(data_store[itr_comb])

            if len(unique_scores) > 1:

                if itr_score == top_three_scores[1]:

                    top_three_store[1].append(data_store[itr_comb])

            if len(unique_scores) > 2:
                
                if itr_score == top_three_scores[2]:

                    top_three_store[2].append(data_store[itr_comb])

        return top_three_store
    

    @staticmethod
    def find_top_three_closest_to_gt(data_store, ground_truth):

        pc_mismatch = []
        pc_match = []
        overall_metric = []
        for itr_comb in data_store:

            nodes = itr_comb['graph']['nodes']
            gt_edges = []
            for itr_gt_edge in ground_truth['edges']:

                if itr_gt_edge[0] in nodes and itr_gt_edge[1] in nodes:

                    gt_edges.append(itr_gt_edge)

            edge_match_count = 0
            edge_mismatch_count = 0
            edges = itr_comb['graph']['edges']
            for itr_edge in edges:

                if itr_edge in gt_edges:

                    edge_match_count += 1

                else:

                    edge_mismatch_count += 1
            
            pc_mismatch.append(100*edge_mismatch_count/len(edges))
            pc_match.append(100*edge_match_count/len(gt_edges))
            overall_metric.append(pc_match[-1] - pc_mismatch[-1])
            pass


        unique_scores = list(set(overall_metric))
        if len(unique_scores) < 3:
            top_score_count = len(unique_scores)
        else:
            top_score_count = 3
        top_three_scores = sorted(unique_scores, reverse=True)[:top_score_count]

        top_three_store = []
        for _ in range(len(top_three_scores)):
            top_three_store.append([])

        for itr_comb, itr_score in enumerate(overall_metric):

            if itr_score == top_three_scores[0]:

                top_three_store[0].append(itr_comb)

            if len(unique_scores) > 1:

                if itr_score == top_three_scores[1]:

                    top_three_store[1].append(itr_comb)

            if len(unique_scores) > 2:
                
                if itr_score == top_three_scores[2]:

                    top_three_store[2].append(itr_comb)

        return top_three_store, overall_metric, pc_match, pc_mismatch

            

    @staticmethod
    def remove_cycle_datapoints(data_store):

        cyclic_graphs = []
        for itr, itr_comb in enumerate(data_store):

            if not itr_comb['graph score']:

                cyclic_graphs.append(itr)

        data_store = [data for itr, data in enumerate(data_store) if itr not in cyclic_graphs] 

        return data_store, cyclic_graphs



    @staticmethod
    def visualise_graph(data_store):
        
        vis_graph = nx.DiGraph()
        plt.figure(figsize=[5, 5])
        # plt.title()
        vis_graph.add_nodes_from(data_store['graph']['nodes'])
        vis_graph.add_edges_from(data_store['graph']['edges'])
        pos = nx.circular_layout(vis_graph)
        nx.draw_networkx_edges(vis_graph, pos, edgelist=data_store['edges cat']['data'], edge_color='red', width=2)
        nx.draw_networkx_edges(vis_graph, pos, edgelist=data_store['edges cat']['knowledge'], edge_color='black', width=2)
        nx.draw_networkx_edges(vis_graph, pos, edgelist=data_store['edges cat']['undirected'], edge_color='blue', width=2)
        nx.draw_networkx_nodes(vis_graph, pos, node_color='black', node_size=500, nodelist=data_store['graph']['nodes'])
        nx.draw_networkx_labels(vis_graph, pos, font_color='green')
        plt.show()

        return vis_graph


    @staticmethod
    def write_pydot(data_store, name):

        vis_graph = analyse_graphs.visualise_graph(data_store)
        vis_graph = nx.drawing.nx_pydot.to_pydot(vis_graph)
        vis_graph.write_png(name+'.png')


    @staticmethod
    def write_data_store(data_store, filepath):

        wb = Workbook()
        for itr_data in data_store:

            wb = analyse_graphs.write_excel_sheet(wb, itr_data, str(itr_data['combination']))
            
        wb.save(filepath)

        
    @staticmethod
    def write_top_three(data_store, filepath):

        wb = Workbook()
        for itr_place, itr_place_graphs in enumerate(data_store, start=1):

            for itr_graph in itr_place_graphs:

                wb = analyse_graphs.write_excel_sheet(wb, itr_graph, str(itr_place) + '--' + str(itr_graph['combination']))

        wb.save(filepath)

 
    @staticmethod
    def write_excel_sheet(wb, data, sheet_name):

        sheet = wb.create_sheet(title=sheet_name)
            
        nodes = data['graph']['nodes']
        knowledge = data['edges cat']['knowledge']
        data_edges = data['edges cat']['data']
        undirected = data['edges cat']['undirected']

        for itr, itr_node in enumerate(nodes, start=1):

            sheet.cell(row=itr, column=1, value=itr_node)

        for itr, itr_kn in enumerate(knowledge, start=1):

            sheet.cell(row=itr, column=2, value=itr_kn[0])
            sheet.cell(row=itr, column=3, value=itr_kn[1])

        for itr, itr_data_edges in enumerate(data_edges, start=1):

            sheet.cell(row=itr, column=4, value=itr_data_edges[0])
            sheet.cell(row=itr, column=5, value=itr_data_edges[1])

        for itr, itr_undi in enumerate(undirected, start=1):

            sheet.cell(row=itr, column=6, value=itr_undi[0])
            sheet.cell(row=itr, column=7, value=itr_undi[1])

        sheet.cell(row=1, column=8, value='graph score')
        sheet.cell(row=2, column=8, value='kn score')
        sheet.cell(row=3, column=8, value='pc max degree')
        sheet.cell(row=1, column=9, value=data['graph score'])
        sheet.cell(row=2, column=9, value=data['knowledge score'])
        sheet.cell(row=3, column=9, value=data[r'% of max degree'])
    
        return wb
    

    @staticmethod
    def read_data_store(filepath):

        wb = load_workbook(filepath)
        sheets = wb.sheetnames

        data_store = []
        for itr_sheet in sheets[1:]:

            nodes = []
            data = []
            knowledge = []
            undi = []

            sheet_obj = wb[itr_sheet]
            for itr_row_num, itr_row in enumerate(sheet_obj.iter_rows(values_only=True)):

                if itr_row_num == 0:

                    graph_score = itr_row[8]
                
                elif itr_row_num == 1:

                    kn_score = itr_row[8]

                elif itr_row_num == 2:

                    pc_max_deg = itr_row[8]  

                if itr_row[0]:
                    nodes.append(itr_row[0])
                if itr_row[1] and itr_row[2]:
                    knowledge.append((itr_row[1], itr_row[2]))
                if itr_row[3] and itr_row[4]:
                    data.append((itr_row[3], itr_row[4]))
                if itr_row[5] and itr_row[6]:
                    undi.append((itr_row[5], itr_row[6]))

            graph_store = {'combination' : itr_sheet,
                           'knowledge score' : kn_score,
                           'data score' : graph_score-kn_score,
                           'graph score' : graph_score,
                           'knowledge count' : len(knowledge),
                           r'% of max degree' : pc_max_deg,
                           'graph' : {'nodes' : nodes,
                                      'edges' : data+knowledge+undi},
                            'edges cat' : {'data' : data,
                                           'knowledge' : knowledge,
                                           'undirected' : undi}}
            
            data_store.append(graph_store)

        return data_store
            

    @staticmethod
    def compile_graphs(top_three_score, data):

        units = list(top_three_score.keys())
        comb = list(product(*(range(2) for _ in units)))

        whole_graph_score = []
        for itr_comb in comb:

            graph_data = {'combination' : itr_comb,
                        'nodes' : [],
                        'edges' : []}
            
            for itr, itr_place in enumerate(itr_comb):

                unit = units[itr]
                unit_graph = top_three_score[unit][itr_place][0]['graph']
                graph_data['nodes'].extend(unit_graph['nodes'])
                graph_data['edges'].extend(unit_graph['edges'])

            # Add flash sensors
            graph_data['nodes'].extend(['V9', 'F_ln'])
            graph_data['edges'].extend([('V9', 'F_ln')])

            graph_data['nodes'] = list(set(graph_data['nodes']))
            graph_data['edges'] = list(set(graph_data['edges'])) 


            cyc_obj = handle_cycles(graph_data, 'hi', 'hi')
            cyc_obj.handle_cycles_ctrl()


            # All loops go through F_rgc so direct all parents to a dummy
            graph_data['nodes'].append('F_rgc_DUMMY')
            for itr_cycle_edge in cyc_obj.cycle_edges:

                if itr_cycle_edge[1] == 'F_rgc':

                    graph_data['edges'].remove(itr_cycle_edge)
                    graph_data['edges'].append((itr_cycle_edge[0], 'F_rgc_DUMMY'))


            
            # Sorting out the F7 <--> P2
            if ('P2', 'F7') in graph_data['edges'] and ('F7', 'P2') in graph_data['edges']:

                P2_parents = []
                F7_parents = []
                for itr_edge in graph_data['edges']:

                    if itr_edge[1] == 'P2':

                        P2_parents.append(itr_edge[0])

                    elif itr_edge[1] == 'F7':

                        F7_parents.append(itr_edge[0])
                        

                trim_P2_parents = [parent for parent in P2_parents if parent != 'F7']
                P2_child_graph = {'nodes' : ['P2']+P2_parents,
                                'edges' : [(parent, 'P2') for parent in P2_parents],
                                'dummy vars' : [],
                                'forbidden' : []}
                fges_obj = run_fges(P2_child_graph, data, 'score kn')
                fges_obj.remove_no_variance()
                P2_score_prior = fges_obj.local_score(['P2'] , P2_parents)
                P2_score_after = fges_obj.local_score(['P2'] , trim_P2_parents)
                P2_penalty = P2_score_prior - P2_score_after

                trim_F7_parents = [parent for parent in F7_parents if parent != 'P2']
                F7_child_graph = {'nodes' : ['F7']+F7_parents,
                                'edges' : [(parent, 'F7') for parent in F7_parents],
                                'dummy vars' : [],
                                'forbidden' : []}
                fges_obj = run_fges(F7_child_graph, data, 'score kn')
                fges_obj.remove_no_variance()
                F7_score_prior = fges_obj.local_score(['F7'] , F7_parents)
                F7_score_after = fges_obj.local_score(['F7'] , trim_F7_parents)
                F7_penalty = F7_score_prior - F7_score_after

                if F7_penalty < P2_penalty:

                    graph_data['edges'].remove(('P2', 'F7'))

                elif F7_penalty > P2_penalty:

                    graph_data['edges'].remove(('F7', 'P2'))

                else:

                    raise ValueError('GIVE UP NOW')

            cyc_obj = handle_cycles(graph_data, 'hi', 'hi')
            cyc_obj.handle_cycles_ctrl()

            if cyc_obj.cycle_edges:

                print(cyc_obj.cycle_edges)

            whole_graph_score.append(graph_data)

        return whole_graph_score


    @staticmethod
    def score_whole_graphs(best_graph_scores, data):

        scores = []
        for itr_graph in best_graph_scores:

            graph = {'nodes' : itr_graph['nodes'],
                    'edges' : itr_graph['edges'],
                    'dummy vars' : [], # there is one but irrelevant at this point
                    'forbidden' : []}
            fges_obj = run_fges(graph, data, 'hi')
            fges_obj.remove_no_variance()
            scores.append(fges_obj.score_dag(itr_graph['edges']))

        for itr, score in enumerate(scores):

            best_graph_scores[itr]['score'] = score

        return best_graph_scores
    

    @staticmethod
    def write_excel_whole_graph(data_store, filepath):

        wb = Workbook()
        for itr_data in data_store:
            
            sheet = wb.create_sheet(title=str(itr_data['combination']))
                
            nodes = itr_data['nodes']
            edges = itr_data['edges']

            for itr, itr_node in enumerate(nodes, start=1):

                sheet.cell(row=itr, column=1, value=itr_node)

            for itr, itr_edge in enumerate(edges, start=1):

                sheet.cell(row=itr, column=2, value=itr_edge[0])
                sheet.cell(row=itr, column=3, value=itr_edge[1])

            sheet.cell(row=1, column=4, value='graph score')
            sheet.cell(row=1, column=5, value=itr_data['score'])

        
        wb.save(filepath)


    @staticmethod
    def read_excel_whole_graph(filepath):

        wb = load_workbook(filepath)
        sheets = wb.sheetnames

        data_store = []
        for itr_sheet in sheets[1:]:

            nodes = []
            edges = []

            sheet_obj = wb[itr_sheet]
            for itr_row_num, itr_row in enumerate(sheet_obj.iter_rows(values_only=True)):

                if itr_row_num == 0:

                    graph_score = itr_row[4]
 

                if itr_row[0]:
                    nodes.append(itr_row[0])
                if itr_row[1] and itr_row[2]:
                    edges.append((itr_row[1], itr_row[2]))

            graph_store = {'combination' : itr_sheet,
                           'score' : graph_score,
                           'nodes' : nodes,
                           'edfes' : edges}

            
            data_store.append(graph_store)

        return data_store


    @staticmethod
    def forbid_within_units(forbid_edges):

        unit_sensors = {'cab' : ['P1', 'F7', 'P2', 'ACAB', 'V6'],
                        'reg' : ['F7', 'P2', 'V6', 'T_reg', 'L_sp', 'P6', 'F_sg', 'T_cyc', 'C_co', 'C_o2', 'V7', 'F_rgc', 'V2', 'F_sc', 'V3'],
                        'react' : ['F_sc', 'V3', 'F_rgc', 'V2', 'F3', 'T2', 'T_r', 'P4'],
                        'furn' : ['F3', 'T1', 'F5', 'V1', 'T3', 'T2'],
                        'distil' : ['T_20', 'T_10', 'T_fra', 'P5', 'F_slurry', 'F_lco', 'V11', 'F_hn', 'V10', 'F_reflux', 'V8'],
                        'wgc' : ['V4', 'AWGC', 'F_lpg'],
                        'flash' : ['V9', 'F_ln']}

        for itr_unit in unit_sensors.values():

            for sensor in itr_unit:
            
                forbid_edges.extend([(sensor, next) for next in itr_unit])
                forbid_edges.extend([(next, sensor) for next in itr_unit])

        return forbid_edges


    @staticmethod
    def forbid_units(forbid_edges):

        unit_sensors = {'cab' : ['P1', 'F7', 'P2', 'ACAB', 'V6'],
                        'reg' : ['F7', 'P2', 'V6', 'T_reg', 'L_sp', 'P6', 'F_sg', 'T_cyc', 'C_co', 'C_o2', 'V7', 'F_rgc', 'V2', 'F_sc', 'V3'],
                        'react' : ['F_sc', 'V3', 'F_rgc', 'V2', 'F3', 'T2', 'T_r', 'P4'],
                        'furn' : ['F3', 'T1', 'F5', 'V1', 'T3', 'T2'],
                        'distil' : ['T_20', 'T_10', 'T_fra', 'P5', 'F_slurry', 'F_lco', 'V11', 'F_hn', 'V10', 'F_reflux', 'V8'],
                        'wgc' : ['V4', 'AWGC', 'F_lpg'],
                        'flash' : ['V9', 'F_ln']}
        
        forbid_pairs = [('reg', 'cab') , ('react', 'cab') , ('furn', 'cab') , ('distil', 'cab') , ('wgc', 'cab') , ('flash', 'cab') , 
                        ('reg', 'furn') , ('react', 'furn') , ('cab', 'furn') , ('distil', 'furn') , ('wgc', 'furn') , ('flash', 'furn') , 
                        ('distil', 'reg') , ('wgc', 'reg') , ('flash', 'reg'),
                        ('distil', 'react') , ('wgc', 'react') , ('flash', 'react') , 
                        ('wgc', 'distil') , ('flash', 'distil') , 
                        ('wgc', 'flash')]
        
        for itr_pair in forbid_pairs:

            unit1_sensors = unit_sensors[itr_pair[0]]
            unit2_sensors = unit_sensors[itr_pair[1]]
            for unit1 in unit1_sensors:

                forbid_edges.extend([(unit1, unit2) for unit2 in unit2_sensors])

        return forbid_edges


    @staticmethod
    def forbid_valves(forbid_edges):

        unit_sensors = {'cab' : ['P1', 'F7', 'P2', 'ACAB', 'V6'],
                'reg' : ['F7', 'P2', 'V6', 'T_reg', 'L_sp', 'P6', 'F_sg', 'T_cyc', 'C_co', 'C_o2', 'V7', 'F_rgc', 'V2', 'F_sc', 'V3'],
                'react' : ['F_sc', 'V3', 'F_rgc', 'V2', 'F3', 'T2', 'T_r', 'P4'],
                'furn' : ['F3', 'T1', 'F5', 'V1', 'T3', 'T2'],
                'distil' : ['T_20', 'T_10', 'T_fra', 'P5', 'F_slurry', 'F_lco', 'V11', 'F_hn', 'V10', 'F_reflux', 'V8'],
                'wgc' : ['V4', 'AWGC', 'F_lpg'],
                'flash' : ['V9', 'F_ln']}
        nodes = []
        for itr in unit_sensors.values():
            nodes += list(itr)
        
        valves = ['V1', 'V2', 'V3', 'V4', 'V6', 'V7', 'V8', 'V9', 'V10', 'V11']

        for itr in valves:

            forbid_edges.extend([(node, itr) for node in nodes])

        return forbid_edges

    

    @staticmethod
    def data_only(best_graph_scores, data, forbid_edges):

        new_data = []
        for itr_data in best_graph_scores:

            graph = {'nodes' : itr_data['nodes'] + ['T_cyc-T_reg'],
                     'edges' : itr_data['edges'],
                    'dummy vars' : [], # there is one but irrelevant at this point
                    'forbidden' : forbid_edges}
            fges_obj = run_fges(graph, data, 'hi')
            fges_obj.run_fges_ctrl()
            added_edges = fges_obj.edges_cat['data'] + fges_obj.edges_cat['undirected']
            new_data.append([added_edges, fges_obj.best_score])

        for itr in range(len(best_graph_scores)):

            best_graph_scores[itr]['data mode'] = new_data[itr]

        return best_graph_scores


    @ staticmethod
    def cluster_mode(best_graph_scores, data, forbid_edges):

        cluster_edges = [('T_r', 'T_20') , ('T_fra', 'F_ln') , ('T_fra', 'F_lpg')]

        new_data = []
        for itr_data in best_graph_scores:

            
            graph = {'nodes' : itr_data['nodes'] + ['T_cyc-T_reg'],
                     'edges' : itr_data['edges'] + cluster_edges,
                    'dummy vars' : [], # there is one but irrelevant at this point
                    'forbidden' : forbid_edges}
            fges_obj = run_fges(graph, data, 'hi')
            fges_obj.run_fges_ctrl()
            added_edges = fges_obj.edges_cat['data'] + fges_obj.edges_cat['undirected']
            new_data.append([added_edges, fges_obj.best_score])

        for itr in range(len(best_graph_scores)):

            best_graph_scores[itr]['cluster mode'] = new_data[itr]
            best_graph_scores[itr]['edges'] += cluster_edges

        return best_graph_scores



    def rigorous_mode(best_graph_scores, data, forbid_edges):

        rigorous_edges = []




    @staticmethod
    def write_cross_unit_graph(data_store, mode, filepath):

        wb = Workbook()
        for itr_data in data_store:
            
            sheet = wb.create_sheet(title=str(itr_data['combination']))
                
            nodes = itr_data['nodes'] + ['T_cyc-T_reg']
            edges = itr_data['edges']

            for itr, itr_node in enumerate(nodes, start=1):

                sheet.cell(row=itr, column=1, value=itr_node)

            for itr, itr_edge in enumerate(edges, start=1):

                sheet.cell(row=itr, column=2, value=itr_edge[0])
                sheet.cell(row=itr, column=3, value=itr_edge[1])

            for itr, itr_add_edge in enumerate(itr_data[mode][0], start=1):

                sheet.cell(row=itr, column=4, value=itr_add_edge[0])
                sheet.cell(row=itr, column=5, value=itr_add_edge[1])

            sheet.cell(row=1, column=6, value='graph score')
            sheet.cell(row=1, column=7, value=itr_data[mode][1])

        
        wb.save(filepath)


    @staticmethod
    def read_cross_unit_graph(filepath):

        wb = load_workbook(filepath)
        sheets = wb.sheetnames

        data_store = []
        for itr_sheet in sheets[1:]:

            nodes = []
            edges = []
            cross_unit = []

            sheet_obj = wb[itr_sheet]
            for itr_row_num, itr_row in enumerate(sheet_obj.iter_rows(values_only=True)):

                if itr_row_num == 0:

                    graph_score = itr_row[6]
 

                if itr_row[0]:
                    nodes.append(itr_row[0])
                if itr_row[1] and itr_row[2]:
                    edges.append((itr_row[1], itr_row[2]))
                if itr_row[3] and itr_row[4]:
                    cross_unit.append((itr_row[3], itr_row[4]))

            graph_store = {'combination' : itr_sheet,
                           'score' : graph_score,
                           'nodes' : nodes,
                           'edges' : edges,
                           'cross unit' : cross_unit}

            
            data_store.append(graph_store)

        return data_store
                    

    @staticmethod
    def compile_knowledge_graphs(top_three_score, data):

        units = list(top_three_score.keys())
        comb = list(product(*(range(2) for _ in units)))

        whole_graph_score = []
        for itr_comb in comb:

            graph_data = {'combination' : itr_comb,
                        'nodes' : [],
                        'edges' : []}
            
            for itr, itr_place in enumerate(itr_comb):

                unit = units[itr]
                unit_graph_list = top_three_score[unit][itr_place]
                kn_len = []
                for itr_graph_list in unit_graph_list:

                    kn_len.append(itr_graph_list['knowledge count'])
                unit_graph = unit_graph_list[kn_len.index(max(kn_len))]['graph']
                graph_data['nodes'].extend(unit_graph['nodes'])
                graph_data['edges'].extend(unit_graph_list[kn_len.index(max(kn_len))]['edges cat']['knowledge'])

            # Add flash sensors
            graph_data['nodes'].extend(['V9', 'F_ln'])
            graph_data['edges'].extend([('V9', 'F_ln')])

            graph_data['nodes'] = list(set(graph_data['nodes']))
            graph_data['edges'] = list(set(graph_data['edges'])) 


            # cyc_obj = handle_cycles(graph_data, 'hi', 'hi')
            # cyc_obj.handle_cycles_ctrl()


            # All loops go through F_rgc so direct all parents to a dummy
            # graph_data['nodes'].append('F_rgc_DUMMY')
            # for itr_cycle_edge in cyc_obj.cycle_edges:

            #     if itr_cycle_edge[1] == 'F_rgc':

            #         graph_data['edges'].remove(itr_cycle_edge)
            #         graph_data['edges'].append((itr_cycle_edge[0], 'F_rgc_DUMMY'))


            
            # Sorting out the F7 <--> P2
            if ('P2', 'F7') in graph_data['edges'] and ('F7', 'P2') in graph_data['edges']:

                P2_parents = []
                F7_parents = []
                for itr_edge in graph_data['edges']:

                    if itr_edge[1] == 'P2':

                        P2_parents.append(itr_edge[0])

                    elif itr_edge[1] == 'F7':

                        F7_parents.append(itr_edge[0])
                        

                trim_P2_parents = [parent for parent in P2_parents if parent != 'F7']
                P2_child_graph = {'nodes' : ['P2']+P2_parents,
                                'edges' : [(parent, 'P2') for parent in P2_parents],
                                'dummy vars' : [],
                                'forbidden' : []}
                fges_obj = run_fges(P2_child_graph, data, 'score kn')
                fges_obj.remove_no_variance()
                P2_score_prior = fges_obj.local_score(['P2'] , P2_parents)
                P2_score_after = fges_obj.local_score(['P2'] , trim_P2_parents)
                P2_penalty = P2_score_prior - P2_score_after

                trim_F7_parents = [parent for parent in F7_parents if parent != 'P2']
                F7_child_graph = {'nodes' : ['F7']+F7_parents,
                                'edges' : [(parent, 'F7') for parent in F7_parents],
                                'dummy vars' : [],
                                'forbidden' : []}
                fges_obj = run_fges(F7_child_graph, data, 'score kn')
                fges_obj.remove_no_variance()
                F7_score_prior = fges_obj.local_score(['F7'] , F7_parents)
                F7_score_after = fges_obj.local_score(['F7'] , trim_F7_parents)
                F7_penalty = F7_score_prior - F7_score_after

                if F7_penalty < P2_penalty:

                    graph_data['edges'].remove(('P2', 'F7'))

                elif F7_penalty > P2_penalty:

                    graph_data['edges'].remove(('F7', 'P2'))

                elif np.isnan(F7_penalty) or np.isnan(P2_penalty):

                    graph_data['edges'].remove(('P2', 'F7'))
                else:

                    raise ValueError('GIVE UP NOW')

            cyc_obj = handle_cycles(graph_data, 'hi', 'hi')
            cyc_obj.handle_cycles_ctrl()

            if cyc_obj.cycle_edges:

                print(cyc_obj.cycle_edges)

            whole_graph_score.append(graph_data)

        return whole_graph_score



    @staticmethod
    def forbid_unit_edges(forbid_edges):

        # backwards forbids
        backwards_store = {}

        node_tiers = {}
        node_tiers['exclude'] = [('F_rgc', 'L_sp') , ('F_sg', 'P6')]
        node_tiers['out'] = ['T_cyc', 'C_co', 'F_sg', 'C_o2', 'F_rgc', 'V7', 'V2']
        node_tiers['unit'] = ['T_reg', 'L_sp', 'P6']
        node_tiers['in'] = ['F7', 'F_sc', 'V6', 'V3', 'P2']
        backwards_store['reg'] = node_tiers

        node_tiers = {}
        node_tiers['exclude'] = [('F_sc', 'T_r')]
        node_tiers['out'] = ['F_sc', 'V3']
        node_tiers['unit'] = ['P4', 'T_r']
        node_tiers['in'] = ['F_rgc', 'V2', 'T2', 'F3']
        backwards_store['react'] = node_tiers

        node_tiers = {}
        node_tiers['exclude'] = [('F_slurry', 'T_20') , ('F_lco', 'T_20') , ('F_hn', 'T_10')]
        node_tiers['out'] = ['F_slurry', 'F_lco', 'F_hn', 'V11', 'V10']
        node_tiers['unit'] = ['T_20', 'T_10', 'T_fra', 'P5']
        node_tiers['in'] = ['F_reflux', 'V8']
        backwards_store['frac'] = node_tiers

        node_tiers = {}
        node_tiers['exclude'] = []
        node_tiers['out'] = ['T2']
        node_tiers['unit'] = ['T3']
        node_tiers['in'] = ['T1', 'F3', 'F5', 'V1']
        backwards_store['furn'] = node_tiers
    
        for unit in backwards_store.values():

            node_tiers_names = ['out', 'unit', 'in']
            for itr_tier in range(len(node_tiers_names)-1):

                curr_tier = node_tiers_names[itr_tier]
                next_tiers = node_tiers_names[itr_tier+1:]
                for itr_next_tier in next_tiers:
                    
                    for itr_tier_node in unit[curr_tier]:

                        forbid_edges.extend([(itr_tier_node, next_tier_node) for next_tier_node in unit[itr_next_tier] if (itr_tier_node, next_tier_node) not in unit['exclude']])

        #cab
        forbid_edges.extend([('ACAB', 'P1') , ('F7', 'P1') , 
                            ('P2', 'P1') , ('V6', 'P1') , ('V6', 'ACAB')])
        

        # forbid diff streams
        #reg
        forbid_edges.extend([('F7', 'F_sc') , ('F_sc', 'F7') , ('P2', 'F_sc') , ('F_sc', 'P2') , ('V6', 'F_sc') , 
                                    ('V3', 'F7') , ('V7', 'F_rgc') , ('V2', 'F_sg') , ('V2', 'T_cyc') , 
                                    ('V2', 'C_co') , ('V2', 'C_o2') ,
                                    ('T_cyc','F_rgc') , ('C_co','F_rgc') , 
                                    ('C_o2', 'F_rgc') , ('F_sg', 'F_rgc') , ('F_rgc', 'T_cyc') , 
                                    ('F_rgc', 'X_co') , ('F_rgc', 'C_o2') , ('F_rgc', 'F_sg')])
        #react
        forbid_edges.extend([('T2', 'F_rgc') , ('F_rgc', 'T2') , ('T2', 'V2') , ('V2', 'T2') , 
                                       ('F3', 'F_rgc') , ('F_rgc', 'F3') , ('F3', 'V2') , ('V2', 'F3')])
        #furn
        forbid_edges.extend([('T1', 'F5') , ('F5', 'T1') , ('F3', 'F5') , ('F5', 'F3') , 
                                       ('V1', 'T1') , ('V1', 'F3')])
        #frac
        forbid_edges.extend([('F_hn', 'F_lco') , ('F_lco', 'F_hn') , 
                                       ('F_hn', 'F_slurry') , ('F_slurry', 'F_hn') ,
                                       ('F_lco', 'F_slurry') , ('F_slurry', 'F_lco') , 
                                       ('V10' , 'F_lco') , ('V10' , 'F_slurry') , 
                                       ('V11' , 'F_hn') , ('V11' , 'F_slurry')])
        
        return forbid_edges
        

    @staticmethod
    def forbid_outlets(forbid_edges):

        unit_sensors = {'cab' : ['P1', 'F7', 'P2', 'ACAB', 'V6'],
                'reg' : ['F7', 'P2', 'V6', 'T_reg', 'L_sp', 'P6', 'F_sg', 'T_cyc', 'C_co', 'C_o2', 'V7', 'F_rgc', 'V2', 'F_sc', 'V3'],
                'react' : ['F_sc', 'V3', 'F_rgc', 'V2', 'F3', 'T2', 'T_r', 'P4'],
                'furn' : ['F3', 'T1', 'F5', 'V1', 'T3', 'T2'],
                'distil' : ['T_20', 'T_10', 'T_fra', 'P5', 'F_slurry', 'F_lco', 'V11', 'F_hn', 'V10', 'F_reflux', 'V8'],
                'wgc' : ['V4', 'AWGC', 'F_lpg'],
                'flash' : ['V9', 'F_ln']}


        forbid_nodes = {'F_sg' : ['react', 'distil', 'flash', 'wgc'],
                        'T_cyc' : ['react', 'distil', 'flash', 'wgc'],
                        'C_co' : ['react', 'distil', 'flash', 'wgc'],
                        'C_o2' : ['react', 'distil', 'flash', 'wgc'],
                        'V7' : ['react', 'distil', 'flash', 'wgc'],
                        'F_hn' : ['flash', 'wgc'],
                        'V10' : ['flash', 'wgc'],
                        'F_lco' : ['flash', 'wgc'],
                        'V11' : ['flash', 'wgc'],
                        'F_slurry' : ['flash', 'wgc'],
                        'F_ln' : ['wgc'],
                        'V9' : ['wgc']}
        
        for itr_node, itr_unit_list in forbid_nodes.items():

            for itr_unit in itr_unit_list:
            
                forbid_edges.extend([(itr_node, node) for node in unit_sensors[itr_unit]])

        return forbid_edges


    @staticmethod
    def compare_cul_modes(modes, all_graphs):

        combs = []
        scores = []
        for itr in all_graphs:

            combs.append(itr['combination'])
            scores.append(itr['score'])

        ordered = sorted(zip(combs, scores), key=lambda x: x[1], reverse=True)
        combs = [sens[0] for sens in ordered]
        scores = [sens[1] for sens in ordered]

        data_store = {}
        for itr in combs:

            data_store[itr] = [None for _ in range(len(modes.keys()))]

        for itr, itr_mode in enumerate(modes.values()):

            for itr_comb in itr_mode:

                combination = itr_comb['combination']
                data_store[combination][itr] = itr_comb['score']

        for itr, itr_mode in enumerate(modes.keys()):

            score = []
            for itr_data in data_store.values():

                score.append(itr_data[itr])

            plt.plot(combs, score, label=list(modes.keys())[itr])

        plt.xlabel('combination')
        plt.ylabel('score')
        plt.legend()
        plt.xticks(rotation='vertical')
        plt.show()


    @staticmethod
    def run_fges_no_knowledge(data, forbid_edges):

        unit_sensors = {'cab' : ['P1', 'F7', 'P2', 'ACAB', 'V6'],
                'reg' : ['F7', 'P2', 'V6', 'T_reg', 'L_sp', 'P6', 'F_sg', 'T_cyc', 'C_co', 'C_o2', 'V7', 'F_rgc', 'V2', 'F_sc', 'V3'],
                'react' : ['F_sc', 'V3', 'F_rgc', 'V2', 'F3', 'T2', 'T_r', 'P4'],
                'furn' : ['F3', 'T1', 'F5', 'V1', 'T3', 'T2'],
                'distil' : ['T_20', 'T_10', 'T_fra', 'P5', 'F_slurry', 'F_lco', 'V11', 'F_hn', 'V10', 'F_reflux', 'V8'],
                'wgc' : ['V4', 'AWGC', 'F_lpg'],
                'flash' : ['V9', 'F_ln']}
        nodes = []
        for itr in unit_sensors.values():
            nodes += list(itr)

        graph = {'nodes' : nodes + ['T_cyc-T_reg'],
                    'edges' : [],
                'dummy vars' : [], # there is one but irrelevant at this point
                'forbidden' : forbid_edges}
        fges_obj = run_fges(graph, data, 'hi')
        fges_obj.run_fges_ctrl()

        return nodes, fges_obj.best_dag, fges_obj.best_score
        

    @staticmethod
    def write_no_knowledge_graph(graph, score, filepath):

        wb = Workbook() 
        sheet = wb.active
        nodes = graph['nodes']
        edges = graph['edges']

        for itr, itr_node in enumerate(nodes, start=1):

            sheet.cell(row=itr, column=1, value=itr_node)

        for itr, itr_edge in enumerate(edges, start=1):

            sheet.cell(row=itr, column=2, value=itr_edge[0])
            sheet.cell(row=itr, column=3, value=itr_edge[1])

        sheet.cell(row=1, column=4, value='graph score')
        sheet.cell(row=1, column=5, value=score)

        wb.save(filepath)


    def get_mode_gt_similarity_score(modes, ground_truth):

        no_var = [('F7', 'P6') , ('P2', 'P6'), ('P6', 'C_co'), ('P6', 'C_o2'), ('F_sg', 'P6') , ('T_reg', 'P6') , ('F3', 'P4'), ('T_r', 'P4'),
                        ('P5', 'F_reflux') , ('F_reflux', 'P5') , ('T_fra', 'P5') , ('P4', 'P5') , 
                        ('P4', 'deltaP') , ('P6', 'deltaP') , ('deltaP', 'F_rgc') , ('deltaP', 'F_sc')]
        bidirectional = [('P2', 'ACAB') , ('T_10', 'T_20') , ('T_fra', 'T_10') , ('F_reflux', 'T_fra')]
        ground_truth_adapt = ground_truth.copy()
        for itr_var in no_var:

            ground_truth_adapt['edges'].remove(itr_var)
        for itr_var in bidirectional:

            ground_truth_adapt['edges'].remove(itr_var)

        new_modes = {}
        for itr_mode, itr_comb in modes.items():

            data_store = itr_comb
            for itr, itr_graph in enumerate(itr_comb):

                # Full gt
                edge_match_count = 0
                edge_mismatch_count = 0
                edges = itr_graph['edges'] + itr_graph['cross unit']
                for itr_edge in edges:

                    if itr_edge in ground_truth['edges']:

                        edge_match_count += 1

                    else:

                        edge_mismatch_count += 1
            

                data_store[itr]['full gt'] = {}
                data_store[itr]['full gt']['pc match'] = 100*edge_match_count/len(ground_truth['edges'])
                data_store[itr]['full gt']['pc mismatch'] = 100*edge_mismatch_count/len(edges)
                data_store[itr]['full gt']['overall metric'] = data_store[itr]['full gt']['pc match'] - data_store[itr]['full gt']['pc mismatch']



                # Adapted gt
                # Removing 0 var and bidirectional
                edge_match_count = 0
                edge_mismatch_count = 0
                edges = itr_graph['edges'] + itr_graph['cross unit']
                for itr_edge in edges:

                    if itr_edge in ground_truth_adapt['edges'] or itr_edge in bidirectional:

                        edge_match_count += 1

                    else:

                        edge_mismatch_count += 1

                data_store[itr]['adapt gt'] = {}
                data_store[itr]['adapt gt']['pc match'] = 100*edge_match_count/len(ground_truth_adapt['edges'])
                data_store[itr]['adapt gt']['pc mismatch'] = 100*edge_mismatch_count/len(edges)
                data_store[itr]['adapt gt']['overall metric'] = data_store[itr]['adapt gt']['pc match'] - data_store[itr]['adapt gt']['pc mismatch']


            new_modes[itr_mode] = data_store

        return new_modes




    def plot_whole_graph(nodes, kn, data):

        edges = kn+data
        vis_graph = nx.DiGraph()
        plt.figure(figsize=[10, 10])
        vis_graph.add_nodes_from(nodes)
        vis_graph.add_edges_from(edges)
        pos = nx.circular_layout(vis_graph)
        nx.draw_networkx_edges(vis_graph, pos, edgelist=data, edge_color='red', width=2)
        nx.draw_networkx_edges(vis_graph, pos, edgelist=kn, edge_color='black', width=2)
        nx.draw_networkx_nodes(vis_graph, pos, node_color='black', node_size=500, nodelist=nodes)
        nx.draw_networkx_labels(vis_graph, pos, font_color='lightblue')
        plt.show()


    @staticmethod
    def compare_gt_similarity(modes, mode, all_graphs):


        combs = []
        scores = []
        for itr in all_graphs:

            combs.append(itr['combination'])
            scores.append(itr['score'])

        ordered = sorted(zip(combs, scores), key=lambda x: x[1], reverse=True)
        combs = [sens[0] for sens in ordered]
        scores = [sens[1] for sens in ordered]


        data_store = {}
        for itr in combs:

            data_store[itr] = [None for _ in range(len(modes.keys()))]

        for itr, itr_mode in enumerate(modes.values()):

            for itr_comb in itr_mode:

                combination = itr_comb['combination']
                data_store[combination][itr] = itr_comb[mode]['pc match']

        for itr, itr_mode in enumerate(modes.keys()):

            score = []
            for itr_data in data_store.values():

                score.append(itr_data[itr])

            plt.plot(combs, score, label=list(modes.keys())[itr])

        plt.xlabel('combination')
        plt.ylabel('pc match')
        plt.legend()
        plt.xticks(rotation='vertical')
        plt.show()



        data_store = {}
        for itr in combs:

            data_store[itr] = [None for _ in range(len(modes.keys()))]

        for itr, itr_mode in enumerate(modes.values()):

            for itr_comb in itr_mode:

                combination = itr_comb['combination']
                data_store[combination][itr] = itr_comb[mode]['pc mismatch']

        for itr, itr_mode in enumerate(modes.keys()):

            score = []
            for itr_data in data_store.values():

                score.append(itr_data[itr])

            plt.plot(combs, score, label=list(modes.keys())[itr])

        plt.xlabel('combination')
        plt.ylabel('pc mismatch')
        plt.legend()
        plt.xticks(rotation='vertical')
        plt.show()



        data_store = {}
        for itr in combs:

            data_store[itr] = [None for _ in range(len(modes.keys()))]

        for itr, itr_mode in enumerate(modes.values()):

            for itr_comb in itr_mode:

                combination = itr_comb['combination']
                data_store[combination][itr] = itr_comb[mode]['overall metric']

        for itr, itr_mode in enumerate(modes.keys()):

            score = []
            for itr_data in data_store.values():

                score.append(itr_data[itr])

            plt.plot(combs, score, label=list(modes.keys())[itr])

        plt.xlabel('combination')
        plt.ylabel('overall metric')
        plt.legend()
        plt.xticks(rotation='vertical')
        plt.show()


        
    @staticmethod
    def score_vs_forbid_edges(data_store):

        node_count = []
        for itr_data in data_store.values():

            node_count.append(len(itr_data[0]['graph']['nodes']))

        forbid_edges = [10, 67, 19, 115, 36, 3]
        max_edges = [node*(node-1) for node in node_count]

        pc_forbid_edges = [100*forbid/maximum for forbid,maximum in zip(forbid_edges,max_edges)]
        ordered = sorted(zip(pc_forbid_edges, list(data_store.keys())), key=lambda x: x[0], reverse=False)
        ordered_scores = [itr[0] for itr in ordered]
        ordered_units = [itr[1] for itr in ordered]

        # score = {}
        # for itr, itr_unit in enumerate(ordered_units):
            
        #     data = data_store[itr_unit]
        #     score[ordered_scores[itr]] = []
        #     for itr_graph in data:

        #         score[ordered_scores[itr]].append(itr_graph['graph score'])

        # plt.boxplot(score.values())


        avgs = []
        for itr_unit in ordered_units:
            
            data = data_store[itr_unit]
            score = []
            for itr_graph in data:

                score.append(itr_graph['graph score'])
            avgs.append(sum(score)/len(score))

        plt.figure()
        plt.plot(ordered_scores, avgs)
        # plt.xticks(range(1, len(ordered_scores)+1) , ordered_scores)
        plt.xlabel('pc of max edges forbidden')
        plt.ylabel('score')
        plt.show()

        plt.figure()
        plt.plot(ordered_units, avgs)
        # plt.xticks(range(1, len(ordered_scores)+1) , ordered_scores)
        plt.xlabel('pc of max edges forbidden')
        plt.ylabel('score')
        plt.show()


    @staticmethod 
    def avg_score_scatt_vs_pc_graph_kn(data_store):

        graph_data = {}
        for itr_unit, itr_data in data_store.items():

            scores = []
            pc_kn = []
            for itr_graph in itr_data:

                scores.append(itr_graph['graph score'])
                pc_kn.append(itr_graph['pc knowledge'])

            percentiles = [stats.percentileofscore(scores, x) for x in scores]
            graph_data[itr_unit] = [percentiles, pc_kn]

        pc_kn = []
        percentiles = []
        for itr_data in graph_data.values():

            percentiles.extend(itr_data[0])
            pc_kn.extend(itr_data[1])

        plt.figure()
        plt.scatter(pc_kn, percentiles)
        plt.xlabel('pc graph kn')
        plt.ylabel('percentile of score in unit data')
        plt.show()


    @staticmethod
    def gt_similarity_score_graph(data_store, ground_truth):

        no_var = {}
        no_var['cab'] = [] 
        no_var['react1'] = [('F7', 'P6'), ('P2', 'P6'), ('P6', 'C_co'), ('P6', 'C_o2'), ('F_sg', 'P6') , ('T_reg', 'P6')] 
        no_var['react2'] = [('F3', 'P4'), ('T_r', 'P4')]
        no_var['distil'] = [('P5', 'F_reflux') , ('F_reflux', 'P5') , ('T_fra', 'P5')]
        no_var['furn'] = []
        no_var['wgc'] = []

        bidirectional = {}
        bidirectional['cab'] = [('P2', 'ACAB')]
        bidirectional['react1'] = []
        bidirectional['react2'] = []
        bidirectional['distil'] = [('T_10', 'T_20') , ('T_fra', 'T_10') , ('F_reflux', 'T_fra')]
        bidirectional['furn'] = []
        bidirectional['wgc'] = []

        add_nodes = {}
        add_nodes['cab'] = []
        add_nodes['react1'] = ['P6']
        add_nodes['react2'] = ['P4']
        add_nodes['distil'] = ['P5']
        add_nodes['furn'] = []
        add_nodes['wgc'] = []
        
        for itr_unit, itr_data in data_store.items():

            # Get full gt for unit
            nodes = itr_data[0]['graph']['nodes'] + add_nodes[itr_unit]
            gt_edges = []
            for itr_gt_edge in ground_truth['edges']:

                if itr_gt_edge[0] in nodes and itr_gt_edge[1] in nodes:

                    gt_edges.append(itr_gt_edge)

            # Get adapt gt for unit
            gt_edges_adapt = gt_edges
            for itr_var in no_var[itr_unit]:

                gt_edges_adapt.remove(itr_var)
            for itr_var in bidirectional[itr_unit]:

                gt_edges_adapt.remove(itr_var)

            unit_data = itr_data
            for itr in range(len(itr_data)):

                # Full gt
                edge_match_count = 0
                edge_mismatch_count = 0
                edges = itr_data[itr]['graph']['edges']
                for itr_edge in edges:

                    if itr_edge in gt_edges:

                        edge_match_count += 1
                    else:
                        edge_mismatch_count += 1
            

                unit_data[itr]['full gt'] = {}
                unit_data[itr]['full gt']['pc match'] = 100*edge_match_count/len(gt_edges)
                unit_data[itr]['full gt']['pc mismatch'] = 100*edge_mismatch_count/len(edges)
                unit_data[itr]['full gt']['overall metric'] = unit_data[itr]['full gt']['pc match'] - unit_data[itr]['full gt']['pc mismatch']


                # Adapted gt
                # Removing 0 var and bidirectional
                edge_match_count = 0
                edge_mismatch_count = 0
                for itr_edge in edges:

                    if itr_edge in gt_edges_adapt or itr_edge in bidirectional[itr_unit]:

                        edge_match_count += 1

                    else:

                        edge_mismatch_count += 1

                unit_data[itr]['adapt gt'] = {}
                unit_data[itr]['adapt gt']['pc match'] = 100*edge_match_count/len(gt_edges_adapt)
                unit_data[itr]['adapt gt']['pc mismatch'] = 100*edge_mismatch_count/len(edges)
                unit_data[itr]['adapt gt']['overall metric'] = unit_data[itr]['adapt gt']['pc match'] - unit_data[itr]['adapt gt']['pc mismatch']


    @staticmethod 
    def avg_sim_score_scatt_vs_pc_graph_kn(data_store, mode):

        graph_data = {}
        for itr_unit, itr_data in data_store.items():

            match = []
            mismatch = []
            overall = []
            pc_kn = []
            for itr_graph in itr_data:

                match.append(itr_graph[mode]['pc match'])
                mismatch.append(itr_graph[mode]['pc mismatch'])
                overall.append(itr_graph[mode]['overall metric'])
                pc_kn.append(itr_graph['pc knowledge'])

            graph_data[itr_unit] = [match, mismatch, overall, pc_kn]

        y_ax = []
        pc_kn = []
        for itr_data in graph_data.values():

            y_ax.extend(itr_data[0])
            pc_kn.extend(itr_data[3])

        plt.figure()
        plt.scatter(pc_kn, y_ax)
        plt.xlabel('pc graph kn')
        plt.ylabel('pc match')
        plt.show()


        y_ax = []
        for itr_data in graph_data.values():

            y_ax.extend(itr_data[1])

        plt.figure()
        plt.scatter(pc_kn, y_ax)
        plt.xlabel('pc graph kn')
        plt.ylabel('pc mismatch')
        plt.show()


        y_ax = []
        for itr_data in graph_data.values():

            y_ax.extend(itr_data[2])

        plt.figure()
        plt.scatter(pc_kn, y_ax)
        plt.xlabel('pc graph kn')
        plt.ylabel('overall metric')
        plt.show()


    @staticmethod
    def get_total_unique_graphs(data_store):

        scores = []
        for itr_data in data_store.values():

            scores.extend([itr_data[itr]['graph score'] for itr in range(len(itr_data))])

        return 100*(len(set(scores)))/len(scores)