
from causallearn.search.ConstraintBased.FCI import fci
from causallearn.utils.PCUtils.BackgroundKnowledge import BackgroundKnowledge
from causallearn.utils.cit import fisherz, kci, d_separation
import matplotlib.pyplot as plt
import networkx as nx
from openpyxl import Workbook




class run_fci_class:

    def __init__(graph_fci, graph):

        graph_fci.data_orig = graph.data
        graph_fci.whole_graph = graph.whole_graph
        graph_fci.var_mapping = graph.var_mapping
        graph_fci.rev_var_mapping = {column: sensor for sensor, column in graph.var_mapping.items()}


    def run_fci_ctrl(graph_fci):

        test = fisherz
        graph_fci.remove_no_variance()
        graph_fci.fci_no_knowledge(test)
        graph_fci.build_background_knowledge()
        graph_fci.fci_with_knowledge(test)
        graph_fci.map_to_sensors()

        trial_name = 'fisherz_cycles_graph1_varyflow_2'
        save_loc = r'C:\Users\byron\OneDrive\Documents\Year 4\CPE440\Final Project\Code Repositiory\Graphs'
        save_path = rf'{save_loc}\{trial_name}.xlsx'
        graph_fci.print_to_excel(save_path)
        # graph_fci.compare_with_knowledge()


    def remove_no_variance(graph_fci):

        column_to_drop = []
        for itr_col in graph_fci.data_orig.columns:

            var_col = graph_fci.data_orig[itr_col].var()
            if var_col < 0.00001:

                column_to_drop.append(itr_col)

        graph_fci.data = graph_fci.data_orig.drop(columns=column_to_drop)
        graph_fci.data_np = graph_fci.data.to_numpy()
        graph_fci.vars = graph_fci.data.columns.tolist()


        dropped_sensors = [graph_fci.rev_var_mapping[column] for column in column_to_drop]
        graph_fci.trim_whole_graph = {}
        graph_fci.trim_whole_graph['nodes'] = [node for node in graph_fci.whole_graph['nodes'] if node not in dropped_sensors]
        graph_fci.trim_whole_graph['edges'] = [(source,dest) for source, dest in graph_fci.whole_graph['edges'] if source not in dropped_sensors and dest not in dropped_sensors]


    def fci_no_knowledge(graph_fci, test):

        g, _ = fci(graph_fci.data_np, independence_test_method=test)
        graph_fci.nodes_obj = g.get_nodes()


    def build_background_knowledge(graph_fci):

        knowledge = BackgroundKnowledge()
        for itr_edge in graph_fci.trim_whole_graph['edges']:

            source_node_var = graph_fci.var_mapping[itr_edge[0]]
            source_node_ind = graph_fci.vars.index(source_node_var)
            # source_node = graph_fci.nodes_obj[source_node_ind]
            source_node = 'X'+str(source_node_ind+1)

            dest_node_var = graph_fci.var_mapping[itr_edge[1]]
            dest_node_ind = graph_fci.vars.index(dest_node_var)
            # dest_node = graph_fci.nodes_obj[dest_node_ind]
            dest_node = 'X'+str(dest_node_ind+1)

            # knowledge.add_required_by_node(source_node, dest_node)
            knowledge.add_required_by_pattern(source_node, dest_node)


        graph_fci.knowledge = knowledge


    def fci_with_knowledge(graph_fci, test):

        _, graph_fci.fci_edges = fci(graph_fci.data_np, independence_test_method=test, background_knowledge=graph_fci.knowledge)
        A=1


    def map_to_sensors(graph_fci):

        graph_fci.fci_edges_tup = []
        for itr_edge in graph_fci.fci_edges:

            source_node = itr_edge.node1.name
            node_num = source_node[1:]
            column_name = graph_fci.vars[int(node_num)-1]
            source_sensor = graph_fci.rev_var_mapping[column_name]

            dest_node = itr_edge.node2.name
            node_num = dest_node[1:]
            column_name = graph_fci.vars[int(node_num)-1]
            dest_sensor = graph_fci.rev_var_mapping[column_name]

            graph_fci.fci_edges_tup.append((source_sensor, dest_sensor))


        print(len(graph_fci.fci_edges_tup))
        print('**')
        vis_graph = nx.DiGraph()
        plt.figure(figsize=[5, 5])
        vis_graph.add_nodes_from(graph_fci.whole_graph['nodes'])
        vis_graph.add_edges_from(graph_fci.fci_edges_tup)
        pos = nx.spring_layout(vis_graph)
        nx.draw(vis_graph, pos, with_labels = True)
        plt.show()
        

        A=1


    def print_to_excel(graph_fci, save_path):

        wb = Workbook()
        ws = wb.active

        for row_index, row_data in enumerate(graph_fci.whole_graph['nodes'], start=1):

            ws.cell(row=row_index, column=1, value=row_data)

        for row_index, row_data in enumerate(graph_fci.fci_edges_tup, start=1):
            for col_index, value in enumerate(row_data, start=2):
                ws.cell(row=row_index, column=col_index, value=value)

        wb.save(save_path)




    # def compare_with_knowledge(graph_fci):

    #     edges_comp = {'equivalent': [],
    #                  'reoriented' : [],
    #                  'additional' : []}
        
    #     for itr_edge in graph_fci.fci_edges_tup:

    #         if itr_edge in graph_fci.trim_whole_graph['edges']:

    #             edges_comp['equivalent'].append(itr_edge)

    #         elif (itr_edge[1], itr_edge[0]) in graph_fci.trim_whole_graph['edges']:

    #             edges_comp['reoriented'].append(itr_edge)

    #         elif itr_edge not in graph_fci.trim_whole_graph['edges']:

    #             edges_comp['additional'].append(itr_edge)

    #     dropped_knowledge = []
    #     for itr_edge in graph_fci.trim_whole_graph['edges']:

    #         if itr_edge not in graph_fci.fci_edges_tup:

    #             dropped_knowledge.append(itr_edge)

        
    #     print(f'Edges still present: {edges_comp["equivalent"]}')
    #     print('*********************')
    #     print(f'Edges reoriented: {edges_comp["reoriented"]}')
    #     print('*********************')
    #     print(f'Edges added: {edges_comp["additional"]}')
    #     print('*********************')
    #     print(f'Edges removed: {dropped_knowledge}')
    #     print('*********************')










